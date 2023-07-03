from openpyxl import Workbook                                       # Biblioteca para crear nuevos archivos de excel
from openpyxl.styles import copy as style_copy                      # Biblioteca para copiar el estilo de las celdas en Excel

################## Agregar cliente ##################

def AgregarCliente(nombreCliente, direccionCliente):
    """
    Esta función permite agregar el nombre y la dirección de un nuevo cliente al archivo de Clientes.
    Además, crea el archivo del nuevo cliente donde se encuentra la información de sus juegos de bloques.
    """
    workbookClientes = load_workbook(filename="Clientes.xlsx", keep_vba=True) #Apertura del archivo de excel de clientes 
    hojaClientes = workbookClientes.active #Hoja del archivo de excel donde están los clientes y su información
    
    i = 3 #Se inicializa el contador para filas en 3 porque en la fila 1 y 2 están los encabezados
    #Ahora se deben recorrer las filas, empezando por la fila 3 para determinar el número de la fila que está libre para incluir un nuevo cliente
    for filaValorNominal in hojaClientes.iter_rows(min_row=3,
                                                    min_col=1,
                                                    max_col=1):
        for celdaValorNominal in filaValorNominal:
            if celdaValorNominal.value == None:
                numFila = i
            else:
                i += 1 

    archivoCliente = Workbook() #Se crea un nuevo archivo de Excel 
    machoteCliente = "./Machotes/Machote para nuevo cliente.xlsm"
    nombreArchivoCliente = nombreCliente + ".xlsx" #El nombre del archivo de Excel va a ser igual al nombre del Cliente
    shutil.copy(machoteCliente, "./Archivos de los clientes/" + nombreArchivoCliente)	

    #Se agrega la información del cliente al archivo de Clientes
    hojaClientes["A"+str(i)] = nombreCliente
    hojaClientes["B"+str(i)] = direccionCliente
    hojaClientes["C"+str(i)] = nombreArchivoCliente

    workbookClientes.save("./Clientes.xlsx")
    return

################## Ingresar juego de bloques/calibrando ##################

def IngresarCalibrando(nombreCliente, objeto, marca, numSerie, material, modelo, grado, unidad):
    archivoCliente = BusquedaClientes(nombreCliente)[2] #Busqueda del archivo del cliente
    workbookCliente = load_workbook(filename="Clientes.xlsx", keep_vba=True)  #Apertura del archivo de excel del cliente

    #Revisar si ya existe algún calibrando registrado con el mismo número de serie
    existeCalibrando = False
    for serieCalibrandoRegistrado in workbookCliente.sheetnames:
        if serieCalibrandoRegistrado == numSerie
            existeCalibrando = True
            break

    if existeCalibrando:
        print("Ya existe un calibrando registrado con el númerio de serie "+ numSerie)

    #Crear una hoja para el nuevo calibrando
    if len(workbookCliente.sheetnames) > 1: #Si ya existen calibrandos registrados en el archivo
        hojaReferencia = workbookCliente.worksheets[0] #Se selecciona la hoja 1 como una referencia para crear la hoja para el nuevo juevo
        hojaNuevoCalibrando = workbookCliente.create_sheet(title = numSerie)
        
        #Se copia la hoja de referencia en la nueva hoja como plantilla
        for row in hojaReferencia.iter_rows():
            for cell in row:
                new_cell = hojaNuevoCalibrando.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell._style = style_copy(cell._style)
    
    else: #Si no se han registrado calibrandos en el archivo del Cliente
        hojaNuevoCalibrando = workbookCliente.worksheets[0]
        hojaNuevoCalibrando.title = numSerie
    
    #Agregar la información del calibrando al archivo del cliente
    hojaNuevoCalibrando["A1"] = "Información del calibrando con identificación " + numSerie
    hojaNuevoCalibrando["C2"] = objeto
    hojaNuevoCalibrando["C3"] = marca
    hojaNuevoCalibrando["C4"] = numSerie
    hojaNuevoCalibrando["C5"] = material #Dropdown con opciones: Acero, cerámica, carburo de tungsteno, carburo de cromo
    hojaNuevoCalibrando["C6"] = modelo
    hojaNuevoCalibrando["C7"] = grado

    hojaNuevoCalibrando["B10"] = "Longitud nominal (" + unidad + ")"

    #Agregar valor nominal e identificación de los bloques del juego
    agregarNuevoBloque = "si"
    numFila = 14 #Se inicia el contador para filas en 14 porque ahí empieza la lista de bloques
    if agregarNuevoBloque == "si":
        valorBloqueIngresar = input("Indique el valor nomial del bloque a ingresar: ") #En la interfaz estos deberían ser campos en una ventana
        idBloqueIngresar = input("Indique la identificación del bloque a ingresar: ")
        #Se agrega la información del bloque a la hoja
        hojaNuevoCalibrando["A"+str(numFila)] = numFila - 13
        hojaNuevoCalibrando["B"+str(numFila)] = valorBloqueIngresar
        hojaNuevoCalibrando["C"+str(numFila)] = idBloqueIngresar

        #Definir el estilo de los bordes de las celdas
        borde_sencillo = Side(border_style = "thin")
        borde_cuadrado = Border(top = borde_sencillo,
                                right = borde_sencillo,
                                bottom = borde_sencillo,
                                left = borde_sencillo)
        
        #Se le da estilo a la nuevas celdas
        hojaNuevoCalibrando["A"+str(numFila)].border = borde_cuadrado
        hojaNuevoCalibrando["B"+str(numFila)].border = borde_cuadrado
        hojaNuevoCalibrando["C"+str(numFila)].border = borde_cuadrado

        numFila += 1
        agregarNuevoBloque = input("¿Desea agregar otro bloque? ")
    else:
        print("Se han ingresado exitosamente los datos del nuevo calibrando.")

    workbookCliente.save(archivoCliente)
    return